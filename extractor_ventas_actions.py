"""
extractor_ventas_actions.py
============================
Versión para GitHub Actions — sin rutas de Windows.
Lee credenciales de variables de entorno (GitHub Secrets).
Se ejecuta en los servidores de GitHub cada día a las 20:15.

Este archivo va en la raíz del repositorio 'comercialespin'.
"""

import imaplib
import email
import email.header
import json
import os
import re
import sys
import logging
import tempfile
import shutil
from datetime import datetime, date, timedelta, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: pip install openpyxl")

# ──────────────────────────────────────────────────────────
#  RUTAS — en Actions el repo está en el directorio actual
# ──────────────────────────────────────────────────────────
REPO_DIR  = Path(".")
JSON_OUT  = REPO_DIR / "data.json"
HTML_SRC  = REPO_DIR / "cuadro_mando_hipopotamo.html"
HTML_DEST = REPO_DIR / "cuadro_mando_hipopotamo.html"  # se sobreescribe
LOG_FILE  = REPO_DIR / "extractor_actions.log"

# ──────────────────────────────────────────────────────────
#  LOGGING
# ──────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────
#  CONFIGURACIÓN — desde variables de entorno (GitHub Secrets)
# ──────────────────────────────────────────────────────────
def load_config() -> dict:
    cfg = {
        "email":              os.environ.get("GMAIL_USER",     ""),
        "password_app":       os.environ.get("GMAIL_PASSWORD", ""),
        "imap_server":        "imap.gmail.com",
        "imap_port":          993,
        "carpeta_busqueda":   "INBOX",
        "asunto_contiene":    os.environ.get("ASUNTO_FILTRO",  "Ventas por caja"),
        "remitente_contiene": os.environ.get("REMITENTE",      "reportes@hipopotamo.com"),
        "nombre_adjunto":     "Ventas por caja.xlsx",
        "buscar_ultimas_horas": 26,
        "hoja_excel":         "VENTAS",
        "vendedor_cavero":    "26",
        "vendedor_ursula":    "61",
    }
    if not cfg["email"] or not cfg["password_app"]:
        log.error("Faltan GMAIL_USER o GMAIL_PASSWORD en los secrets de GitHub.")
        log.error("Ve a: github.com/civcomercial2010-cmyk/comercialespin/settings/secrets/actions")
        sys.exit(1)
    return cfg

# ──────────────────────────────────────────────────────────
#  IMAP — igual que el script local
# ──────────────────────────────────────────────────────────
def decode_header_value(raw) -> str:
    if raw is None: return ""
    parts = email.header.decode_header(raw)
    decoded = []
    for chunk, enc in parts:
        if isinstance(chunk, bytes):
            decoded.append(chunk.decode(enc or "utf-8", errors="replace"))
        else:
            decoded.append(chunk)
    return " ".join(decoded)

def connect_gmail(cfg: dict) -> imaplib.IMAP4_SSL:
    log.info(f"Conectando a {cfg['imap_server']}:993 como {cfg['email']}")
    conn = imaplib.IMAP4_SSL(cfg["imap_server"], int(cfg.get("imap_port", 993)))
    conn.login(cfg["email"], cfg["password_app"])
    log.info("Conexión IMAP establecida.")
    return conn

def find_latest_email(conn, cfg) -> bytes | None:
    carpeta = cfg.get("carpeta_busqueda", "INBOX")
    conn.select(carpeta, readonly=True)
    criterios = []
    if cfg.get("asunto_contiene"):
        criterios.append(f'SUBJECT "{cfg["asunto_contiene"]}"')
    if cfg.get("remitente_contiene"):
        criterios.append(f'FROM "{cfg["remitente_contiene"]}"')
    horas = int(cfg.get("buscar_ultimas_horas", 26))
    desde = (datetime.now(timezone.utc) - timedelta(hours=horas)).strftime("%d-%b-%Y")
    criterios.append(f'SINCE {desde}')
    search_str = "(" + " ".join(criterios) + ")" if len(criterios) > 1 else criterios[0]
    log.info(f"Buscando: {search_str}")
    _, data = conn.search(None, search_str)
    ids = data[0].split()
    if not ids:
        log.warning("No se encontraron emails.")
        return None
    nombre_cfg = cfg.get("nombre_adjunto", "").lower().replace(".xlmx", ".xlsx")
    for uid in reversed(ids):
        _, full = conn.fetch(uid, "(RFC822)")
        msg = email.message_from_bytes(full[0][1])
        for part in msg.walk():
            fn_raw = part.get_filename()
            if not fn_raw: continue
            fn = decode_header_value(fn_raw).strip().lower().replace(".xlmx", ".xlsx")
            if not nombre_cfg or nombre_cfg in fn or fn in nombre_cfg:
                log.info(f"Email encontrado UID {uid.decode()} con adjunto '{fn}'")
                return uid
    log.warning("Ningún email con el adjunto esperado.")
    return None

def download_attachment(conn, uid, cfg, tmp_dir) -> Path | None:
    _, msg_data = conn.fetch(uid, "(RFC822)")
    msg = email.message_from_bytes(msg_data[0][1])
    nombre_cfg = cfg.get("nombre_adjunto", "").lower().replace(".xlmx", ".xlsx")
    for part in msg.walk():
        fn_raw = part.get_filename()
        if not fn_raw: continue
        fn = decode_header_value(fn_raw).strip()
        fn_norm = fn.lower().replace(".xlmx", ".xlsx")
        if not fn_norm.endswith(".xlsx"): continue
        if nombre_cfg and nombre_cfg not in fn_norm and fn_norm not in nombre_cfg: continue
        dest = Path(tmp_dir) / fn_norm
        dest.write_bytes(part.get_payload(decode=True))
        log.info(f"Adjunto descargado: {fn}")
        return dest
    log.error("No se encontró adjunto .xlsx.")
    return None

# ──────────────────────────────────────────────────────────
#  PARSEO DEL EXCEL — estructura real confirmada
# ──────────────────────────────────────────────────────────
def parse_amount(cell_value) -> float | None:
    if cell_value is None: return None
    if isinstance(cell_value, (int, float)): return float(cell_value)
    s = str(cell_value).strip().replace("\u00a0","").replace(" ","")
    if re.match(r"^-?[\d.]+,\d+$", s):
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", "")
    try: return float(s)
    except: return None

def detect_fecha_hasta(ws) -> date | None:
    pattern = re.compile(
        r"Fecha\s+hasta[:\s]+(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})",
        re.IGNORECASE
    )
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        for cell in row:
            if cell is None: continue
            m = pattern.search(str(cell))
            if m:
                d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
                if y < 100: y += 2000
                try: return date(y, mo, d)
                except: pass
    return None

def comercial_month_from_date(fecha_hasta: date) -> tuple[int, int]:
    if fecha_hasta.day <= 25:
        return fecha_hasta.year, fecha_hasta.month
    if fecha_hasta.month == 12:
        return fecha_hasta.year + 1, 1
    return fecha_hasta.year, fecha_hasta.month + 1

def extract_data(xlsx_path: Path, cfg: dict) -> dict:
    log.info(f"Procesando: {xlsx_path.name}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_name = cfg.get("hoja_excel", "VENTAS")
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
        log.warning(f"Usando hoja: '{sheet_name}'")
    ws = wb[sheet_name]

    fecha_hasta = detect_fecha_hasta(ws)
    if not fecha_hasta:
        log.error("No se detectó 'Fecha hasta'.")
        wb.close(); return {}
    log.info(f"Fecha hasta: {fecha_hasta.strftime('%d/%m/%Y')}")
    cm_year, cm_month = comercial_month_from_date(fecha_hasta)
    log.info(f"Mes comercial: {cm_year}-{cm_month:02d}")

    # Cabecera de vendedores
    header_row = -1
    col_grupo = 3; col_importe = 4; col_consumo_t = 7; col_alta_t = 8

    for idx, row in enumerate(ws.iter_rows(max_row=120, values_only=True), start=1):
        r = [str(c).strip().lower() if c is not None else '' for c in row]
        if not r: continue
        if r[0] == 'vendedor' and 'grupo' in r:
            header_row    = idx
            col_grupo     = r.index('grupo') + 1
            col_importe   = next((i+1 for i,v in enumerate(r) if 'base imponible' in v), 4)
            col_consumo_t = next((i+1 for i,v in enumerate(r) if 'consumo' in v and 'tarj' in v), 7)
            col_alta_t    = next((i+1 for i,v in enumerate(r) if 'alta' in v and 'tarj' in v), 8)
            log.info(f"Cabecera fila {idx}: grupo=col{col_grupo} imp=col{col_importe} "
                     f"consumo_t=col{col_consumo_t} alta_t=col{col_alta_t}")
            break

    vend_c = str(cfg.get("vendedor_cavero", "26"))
    vend_u = str(cfg.get("vendedor_ursula", "61"))

    cav_prof = cav_dist = cav_ind = 0.0
    cav_bonos_altas = cav_bonos_consumo = None
    urs_total = urs_bonos_altas = urs_bonos_consumo = None

    def get_float(row, col):
        if col <= 0 or col > len(row): return 0.0
        v = row[col - 1]
        if isinstance(v, (int, float)): return float(v)
        try: return float(str(v).strip().replace(',','.'))
        except: return 0.0

    data_start = header_row + 1 if header_row > 0 else 96
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not any(c is not None for c in row): continue
        c0 = str(row[0]).strip() if row[0] is not None else ''
        if not c0: continue

        grupo   = str(row[col_grupo-1]).strip().upper() if col_grupo <= len(row) and row[col_grupo-1] else ''
        importe = get_float(row, col_importe)

        if c0 == vend_c:
            if   'PROFESIONAL' in grupo:                        cav_prof += importe
            elif 'DISTRIBU'    in grupo or 'DIS V2' == grupo:  cav_dist += importe
            elif 'INDUSTRIA'   in grupo:                        cav_ind  += importe
            continue

        if re.match(r'total\s+vendedor\s+' + re.escape(vend_c) + r'\b', c0, re.IGNORECASE):
            ct = get_float(row, col_consumo_t); at = get_float(row, col_alta_t)
            if at: cav_bonos_altas   = abs(at)
            if ct: cav_bonos_consumo = abs(ct)
            log.info(f"Total Cavero: bonos altas={at:.2f} consumo={ct:.2f}")
            continue

        if re.match(r'total\s+vendedor\s+' + re.escape(vend_u) + r'\b', c0, re.IGNORECASE):
            urs_total = importe
            ct = get_float(row, col_consumo_t); at = get_float(row, col_alta_t)
            if at: urs_bonos_altas   = abs(at)
            if ct: urs_bonos_consumo = abs(ct)
            log.info(f"Total Úrsula: ventas={importe:.2f} bonos altas={at:.2f} consumo={ct:.2f}")
            continue

    wb.close()

    log.info(f"Cavero Prof={cav_prof:.2f} Dist={cav_dist:.2f} Ind={cav_ind:.2f}")
    log.info(f"Úrsula Total={urs_total}")

    return {
        "fecha_hasta":        fecha_hasta.isoformat(),
        "cm_year":            cm_year,
        "cm_month":           cm_month,
        "cavero_prof":        round(cav_prof, 2) if cav_prof else None,
        "cavero_dist":        round(cav_dist, 2) if cav_dist else None,
        "cavero_ind":         round(cav_ind,  2) if cav_ind  else None,
        "cavero_bonos_altas": round(cav_bonos_altas,   2) if cav_bonos_altas   else None,
        "cavero_bonos_cons":  round(cav_bonos_consumo, 2) if cav_bonos_consumo else None,
        "ursula":             round(urs_total, 2) if urs_total else None,
        "ursula_bonos_altas": round(urs_bonos_altas,   2) if urs_bonos_altas   else None,
        "ursula_bonos_cons":  round(urs_bonos_consumo, 2) if urs_bonos_consumo else None,
    }

# ──────────────────────────────────────────────────────────
#  ACTUALIZAR data.json
# ──────────────────────────────────────────────────────────
def update_json(result: dict):
    existing = {}
    if JSON_OUT.exists():
        try:
            with open(JSON_OUT, encoding="utf-8") as f:
                existing = json.load(f)
        except Exception as e:
            log.warning(f"data.json inválido: {e}")

    yr  = result["cm_year"]
    mon = result["cm_month"] - 1
    key = f"{yr}-{result['cm_month']:02d}"

    existing.setdefault("cavero", {})
    existing.setdefault("ursula", {})
    existing["cavero"].setdefault("monthly",       {})
    existing["cavero"].setdefault("industria",      {})
    existing["cavero"].setdefault("distribuidores", {})
    existing["cavero"].setdefault("bonos",          {})
    existing["ursula"].setdefault("monthly",        {})
    existing["ursula"].setdefault("bonos",          {})

    existing["cavero"]["monthly"].setdefault(str(yr), [None]*12)
    existing["ursula"]["monthly"].setdefault(str(yr), [None]*12)

    if result.get("cavero_prof") is not None:
        existing["cavero"]["monthly"][str(yr)][mon]  = result["cavero_prof"]
    if result.get("cavero_ind")  is not None:
        existing["cavero"]["industria"][key]          = result["cavero_ind"]
    if result.get("cavero_dist") is not None:
        existing["cavero"]["distribuidores"][key]     = result["cavero_dist"]
    if result.get("ursula")      is not None:
        existing["ursula"]["monthly"][str(yr)][mon]   = result["ursula"]

    existing["cavero"]["bonos"].setdefault(key, {})
    existing["ursula"]["bonos"].setdefault(key, {})
    if result.get("cavero_bonos_altas") is not None:
        existing["cavero"]["bonos"][key]["altas"]   = result["cavero_bonos_altas"]
    if result.get("cavero_bonos_cons")  is not None:
        existing["cavero"]["bonos"][key]["consumo"] = result["cavero_bonos_cons"]
    if result.get("ursula_bonos_altas") is not None:
        existing["ursula"]["bonos"][key]["altas"]   = result["ursula_bonos_altas"]
    if result.get("ursula_bonos_cons")  is not None:
        existing["ursula"]["bonos"][key]["consumo"] = result["ursula_bonos_cons"]

    existing["lastLoadDate"] = result["fecha_hasta"]
    existing["lastRunTs"]    = datetime.now().isoformat()

    with open(JSON_OUT, "w", encoding="utf-8") as f:
        json.dump(existing, f, ensure_ascii=False, indent=2)
    log.info(f"data.json actualizado → {JSON_OUT}")

# ──────────────────────────────────────────────────────────
#  INYECTAR DATOS EN EL HTML
# ──────────────────────────────────────────────────────────
_DATA_MARKER = "/* __DATOS_INYECTADOS__ */"

def build_html(data_payload: dict):
    if not HTML_SRC.exists():
        log.warning(f"HTML no encontrado: {HTML_SRC}")
        return
    html = HTML_SRC.read_text(encoding="utf-8")
    if _DATA_MARKER not in html:
        log.warning("Marcador __DATOS_INYECTADOS__ no encontrado en el HTML.")
        return

    payload_js = json.dumps(data_payload, ensure_ascii=False)
    lines = [
        "/* __DATOS_INYECTADOS__ */",
        "  (function(){",
        "    var _d = " + payload_js + ";",
        "    try {",
        "      var ls = JSON.parse(localStorage.getItem('hipopotamo_v2') || '{}');",
        "      ['cavero','ursula'].forEach(function(p){",
        "        if(!_d[p]) return;",
        "        ls[p] = ls[p] || {};",
        "        if(_d[p].monthly){",
        "          ls[p].monthly = ls[p].monthly || {};",
        "          Object.entries(_d[p].monthly).forEach(function(e){",
        "            var yr=e[0], arr=e[1];",
        "            ls[p].monthly[yr] = ls[p].monthly[yr] || Array(12).fill(null);",
        "            arr.forEach(function(v,i){ if(v!==null) ls[p].monthly[yr][i]=v; });",
        "          });",
        "        }",
        "        if(_d[p].industria)      ls[p].industria      = Object.assign(ls[p].industria||{},      _d[p].industria);",
        "        if(_d[p].distribuidores) ls[p].distribuidores = Object.assign(ls[p].distribuidores||{}, _d[p].distribuidores);",
        "        if(_d[p].bonos)          ls[p].bonos          = Object.assign(ls[p].bonos||{},          _d[p].bonos);",
        "      });",
        "      if(_d.lastLoadDate){",
        "        var rd=new Date(_d.lastLoadDate), ld=ls.lastLoadDate?new Date(ls.lastLoadDate):null;",
        "        if(!ld||rd>ld) ls.lastLoadDate=_d.lastLoadDate;",
        "      }",
        "      ls.lastRunTs = _d.lastRunTs || new Date().toISOString();",
        "      localStorage.setItem('hipopotamo_v2', JSON.stringify(ls));",
        "    } catch(e) { console.warn('Inyeccion datos:', e); }",
        "  })();",
    ]
    injection = chr(10).join(lines)
    HTML_DEST.write_text(html.replace(_DATA_MARKER, injection, 1), encoding="utf-8")
    log.info(f"HTML actualizado con datos inyectados.")

# ──────────────────────────────────────────────────────────
#  MAIN
# ──────────────────────────────────────────────────────────
def main():
    log.info("=" * 60)
    log.info("Extractor Ventas — GitHub Actions")
    log.info(f"Hora UTC: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')}")
    log.info("=" * 60)

    cfg  = load_config()
    conn = connect_gmail(cfg)
    uid  = find_latest_email(conn, cfg)

    if uid is None:
        log.warning("Sin email válido. Nada que actualizar.")
        conn.logout()
        sys.exit(0)

    with tempfile.TemporaryDirectory() as tmp:
        xlsx_path = download_attachment(conn, uid, cfg, tmp)
        conn.logout()
        if xlsx_path is None:
            sys.exit(1)
        result = extract_data(xlsx_path, cfg)

    if not result:
        log.error("Extracción fallida.")
        sys.exit(1)

    # Construir payload con data.json existente + datos nuevos
    data_payload = {}
    if JSON_OUT.exists():
        try:
            with open(JSON_OUT, encoding="utf-8") as f:
                data_payload = json.load(f)
        except Exception:
            pass

    yr  = result["cm_year"]
    mon = result["cm_month"] - 1
    key = f"{yr}-{result['cm_month']:02d}"
    data_payload.setdefault("cavero", {})
    data_payload.setdefault("ursula", {})
    data_payload["cavero"].setdefault("monthly",       {})
    data_payload["cavero"].setdefault("industria",      {})
    data_payload["cavero"].setdefault("distribuidores", {})
    data_payload["cavero"].setdefault("bonos",          {})
    data_payload["ursula"].setdefault("monthly",        {})
    data_payload["ursula"].setdefault("bonos",          {})
    data_payload["cavero"]["monthly"].setdefault(str(yr), [None]*12)
    data_payload["ursula"]["monthly"].setdefault(str(yr), [None]*12)

    if result.get("cavero_prof") is not None:
        data_payload["cavero"]["monthly"][str(yr)][mon]  = result["cavero_prof"]
    if result.get("cavero_ind")  is not None:
        data_payload["cavero"]["industria"][key]          = result["cavero_ind"]
    if result.get("cavero_dist") is not None:
        data_payload["cavero"]["distribuidores"][key]     = result["cavero_dist"]
    if result.get("ursula")      is not None:
        data_payload["ursula"]["monthly"][str(yr)][mon]   = result["ursula"]
    data_payload["cavero"]["bonos"].setdefault(key, {})
    data_payload["ursula"]["bonos"].setdefault(key, {})
    if result.get("cavero_bonos_altas"): data_payload["cavero"]["bonos"][key]["altas"]   = result["cavero_bonos_altas"]
    if result.get("cavero_bonos_cons"):  data_payload["cavero"]["bonos"][key]["consumo"] = result["cavero_bonos_cons"]
    if result.get("ursula_bonos_altas"): data_payload["ursula"]["bonos"][key]["altas"]   = result["ursula_bonos_altas"]
    if result.get("ursula_bonos_cons"):  data_payload["ursula"]["bonos"][key]["consumo"] = result["ursula_bonos_cons"]
    data_payload["lastLoadDate"] = result["fecha_hasta"]
    data_payload["lastRunTs"]    = datetime.now(timezone.utc).isoformat()

    build_html(data_payload)
    update_json(result)

    log.info("Proceso completado. GitHub Actions hará el commit y push.")
    log.info("=" * 60)

if __name__ == "__main__":
    main()
